"""
GUIRRA SOLUTION — Pipeline Principal de Conciliação
Executa todos os 11 módulos e exporta o Excel consolidado com formatação profissional
"""

import pandas as pd
import numpy as np
import os, sys, warnings
from datetime import datetime
warnings.filterwarnings("ignore")

sys.path.insert(0, os.path.dirname(__file__))
from utils import (get_logger, carregar_base_interna, carregar_base_interna_debitos,
                   carregar_third_party, carregar_debitos_third_party)
from m00_dicionario_custos import (get_df_mdr, get_df_tarifas, get_df_impostos,
                                    get_df_cb_reasons)
from m01_pagamentos_anulacoes import executar as m01
from m02_m11_modulos import (m02_reembolsos, m03_chargebacks_notificados,
                              m04_disputas, m05_chargebacks_debitados,
                              m06_taxas_custos, m07_fluxo_caixa,
                              m08_parcelamento, m09_antecipacoes,
                              m10_recebiveis_liquidos, m11_remessas_comercios)

try:
    from openpyxl import load_workbook
    from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.chart.series import DataPoint
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

log = get_logger("guirra.pipeline")

# Pasta outputs/ na raiz do projeto — criada automaticamente
_ROOT = os.path.join(os.path.dirname(__file__), "..", "..")
OUTPUT_DIR = os.path.join(_ROOT, "outputs")
os.makedirs(OUTPUT_DIR, exist_ok=True)
OUTPUT_PATH = os.path.join(OUTPUT_DIR, "GuirraSolution_Fase2_Conciliacao.xlsx")

# ─── CORES GUIRRA SOLUTION ───────────────────────────────────────────────────
C = {
    "dark":    "1A2B4A",
    "mid":     "1E5799",
    "acc":     "00A878",
    "red":     "E63946",
    "yellow":  "FFB703",
    "light":   "EEF4FB",
    "gray":    "F5F7FA",
    "white":   "FFFFFF",
    "border":  "C9D6E8",
    "green":   "D4EDDA",
    "red_bg":  "F8D7DA",
    "yel_bg":  "FFF3CD",
}

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=11, italic=False):
    return Font(bold=bold, color=color, size=size, name="Arial", italic=italic)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border():
    s = Side(style="thin", color=C["border"])
    return Border(left=s, right=s, top=s, bottom=s)

def style_header(ws, row, ncols, start_col=1):
    for c in range(start_col, start_col + ncols):
        cell = ws.cell(row=row, column=c)
        cell.fill      = fill(C["dark"])
        cell.font      = font(bold=True, color=C["white"], size=10)
        cell.alignment = align("center")
        cell.border    = thin_border()

def style_data_row(ws, row, ncols, start_col=1, even=True):
    bg = C["gray"] if not even else C["white"]
    for c in range(start_col, start_col + ncols):
        cell = ws.cell(row=row, column=c)
        cell.fill      = fill(bg)
        cell.font      = font(size=9)
        cell.alignment = align()
        cell.border    = thin_border()

def style_status_cell(cell, status):
    mapping = {
        "CONCILIADO":        (C["green"],  C["acc"],   True),
        "DIVERGENTE":        (C["red_bg"], C["red"],   True),
        "PENDENTE":          (C["yel_bg"], C["yellow"],True),
        "SEM_PAR":           (C["red_bg"], C["red"],   True),
        "EXCLUIDO_NEGADA":   (C["gray"],   "888888",   False),
        "EXCLUIDO_EXPIRADA": (C["gray"],   "888888",   False),
        "ANOMALIA":          (C["red_bg"], C["red"],   True),
    }
    if status in mapping:
        bg, fg, bold = mapping[status]
        cell.fill = fill(bg)
        cell.font = font(bold=bold, color=fg, size=9)
        cell.alignment = align("center")

def write_df_to_sheet(ws, df, start_row=2, max_rows=50000):
    """Escreve DataFrame formatado na planilha."""
    cols = list(df.columns)
    # Header
    for ci, col in enumerate(cols, 1):
        ws.cell(row=start_row, column=ci).value = col
    style_header(ws, start_row, len(cols))

    # Dados
    df_out = df.head(max_rows).copy()
    for col in df_out.select_dtypes(include=["datetimetz"]).columns:
        df_out[col] = df_out[col].dt.strftime("%Y-%m-%d %H:%M")
    for col in df_out.select_dtypes(include=["datetime64[ns]"]).columns:
        df_out[col] = df_out[col].dt.strftime("%Y-%m-%d")

    for ri, row_data in enumerate(df_out.itertuples(index=False), start_row + 1):
        even = (ri % 2 == 0)
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci)
            cell.value = None if (isinstance(val, float) and np.isnan(val)) else val
            cell.font      = font(size=9)
            cell.alignment = align()
            cell.border    = thin_border()
            cell.fill      = fill(C["white"] if even else C["gray"])
            # Colorir coluna de status
            if cols[ci-1] == "STATUS_CONCILIACAO":
                style_status_cell(cell, str(val))
            # Valores monetários
            if cols[ci-1].startswith("VALOR") or cols[ci-1] in ["DIFF_VALOR","DIFF_DEBITO","CUSTO_TOTAL"]:
                if isinstance(val, (int, float)) and not (isinstance(val, float) and np.isnan(val)):
                    cell.number_format = 'R$ #,##0.00'
            if "PCT" in cols[ci-1] or "MDR" in cols[ci-1] and "PCT" in cols[ci-1]:
                if isinstance(val, (int, float)) and not (isinstance(val, float) and np.isnan(val)):
                    cell.number_format = '0.00"%"'

    # Auto-width
    for ci, col in enumerate(cols, 1):
        max_len = max(len(str(col)), 8)
        for ri in range(start_row, start_row + min(len(df_out), 100) + 1):
            val = ws.cell(row=ri, column=ci).value
            if val:
                max_len = max(max_len, min(len(str(val)), 40))
        ws.column_dimensions[get_column_letter(ci)].width = max_len + 2

    return start_row + len(df_out) + 1

def create_tab_header(ws, title, subtitle="", row=1):
    """Cabeçalho azul escuro da aba."""
    ws.row_dimensions[row].height = 40
    cell = ws.cell(row=row, column=1, value=f"  GUIRRA SOLUTION  |  {title}")
    cell.fill      = fill(C["dark"])
    cell.font      = Font(bold=True, color=C["acc"], size=14, name="Arial")
    cell.alignment = align("left", "center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    if subtitle:
        ws.row_dimensions[row+1].height = 22
        cell2 = ws.cell(row=row+1, column=1, value=f"  {subtitle}")
        cell2.fill = fill(C["mid"])
        cell2.font = Font(color=C["white"], size=10, name="Arial")
        cell2.alignment = align("left", "center")
        ws.merge_cells(start_row=row+1, start_column=1, end_row=row+1, end_column=12)

def create_kpi_row(ws, kpis: list, start_row=3):
    """kpis: list of (label, value, color)"""
    ws.row_dimensions[start_row].height = 52
    for i, (label, value, color) in enumerate(kpis):
        col = i * 3 + 1
        ws.merge_cells(start_row=start_row, start_column=col,
                       end_row=start_row, end_column=col+2)
        cell = ws.cell(row=start_row, column=col)
        cell.value     = f"{value}\n{label}"
        cell.fill      = fill(C["dark"])
        cell.font      = Font(bold=True, color=color, size=14, name="Arial")
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)


def run_pipeline():
    log.info("=" * 70)
    log.info("GUIRRA SOLUTION — Pipeline de Conciliação Financeira")
    log.info(f"Início: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    log.info("=" * 70)

    # ── Carregar bases uma vez ────────────────────────────────────────────────
    log.info("Carregando bases de dados...")
    df_bi  = carregar_base_interna(deduplicar_nsu=True)
    df_bid = carregar_base_interna_debitos()
    df_tp  = carregar_third_party(agrupar_por_nsu=True)
    df_tp_raw = carregar_third_party(agrupar_por_nsu=False)
    df_dtp = carregar_debitos_third_party()

    # ── Executar todos os módulos ─────────────────────────────────────────────
    log.info("Executando módulos de conciliação...")
    r01 = m01(df_bi, df_tp)
    r02 = m02_reembolsos(df_bid, df_dtp)
    r03 = m03_chargebacks_notificados(df_bi, df_tp_raw)
    r04 = m04_disputas(df_bi)
    r05 = m05_chargebacks_debitados(df_bi, df_tp_raw)
    r06 = m06_taxas_custos(df_tp)
    r07 = m07_fluxo_caixa(df_bi, df_bid, df_tp_raw)
    r08 = m08_parcelamento(df_bi, df_tp_raw)
    r09 = m09_antecipacoes(df_tp_raw)
    r10 = m10_recebiveis_liquidos(df_tp, df_bi)
    r11 = m11_remessas_comercios(df_tp_raw)

    all_kpis = {
        "M01": r01["kpis"], "M02": r02["kpis"], "M03": r03["kpis"],
        "M04": r04["kpis"], "M05": r05["kpis"], "M06": r06["kpis"],
        "M07": r07["kpis"], "M08": r08["kpis"], "M09": r09["kpis"],
        "M10": r10["kpis"], "M11": r11["kpis"],
    }

    log.info("Gerando Excel consolidado...")

    # ── Criar Excel com openpyxl ──────────────────────────────────────────────
    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:

        # ── ABA SUMÁRIO ───────────────────────────────────────────────────────
        df_sumario = pd.DataFrame([
            {"MÓDULO": "M01 – Pagamentos e Anulações",
             "TOTAL": r01["kpis"]["total"],
             "CONCILIADOS": r01["kpis"]["conciliados"],
             "DIVERGENTES": r01["kpis"]["divergentes"],
             "SEM_PAR": r01["kpis"]["sem_par"],
             "TAXA_CONCILIAÇÃO_%": r01["kpis"]["taxa_conciliacao_pct"],
             "VALOR_CONCILIADO_R$": r01["kpis"]["valor_conciliado"]},
            {"MÓDULO": "M02 – Reembolsos",
             "TOTAL": r02["kpis"]["total"],
             "CONCILIADOS": r02["kpis"]["conciliados"],
             "DIVERGENTES": r02["kpis"]["divergentes"],
             "SEM_PAR": r02["kpis"]["sem_par"],
             "TAXA_CONCILIAÇÃO_%": r02["kpis"]["taxa_conciliacao_pct"],
             "VALOR_CONCILIADO_R$": r02["kpis"]["valor_reembolsado"]},
            {"MÓDULO": "M03 – CB Notificados",
             "TOTAL": r03["kpis"]["total"],
             "CONCILIADOS": r03["kpis"]["conciliados"],
             "DIVERGENTES": r03["kpis"]["divergentes"],
             "SEM_PAR": r03["kpis"]["sem_par"],
             "TAXA_CONCILIAÇÃO_%": r03["kpis"]["taxa_conciliacao_pct"],
             "VALOR_CONCILIADO_R$": r03["kpis"]["valor_total_cb"]},
            {"MÓDULO": "M04 – Disputas",
             "TOTAL": r04["kpis"]["total_cb"] + r04["kpis"]["total_repr"],
             "CONCILIADOS": r04["kpis"]["total_repr"],
             "DIVERGENTES": 0,
             "SEM_PAR": r04["kpis"]["total_cb"],
             "TAXA_CONCILIAÇÃO_%": round(r04["kpis"]["total_repr"]/(r04["kpis"]["total_cb"]+r04["kpis"]["total_repr"])*100,2) if r04["kpis"]["total_cb"]+r04["kpis"]["total_repr"]>0 else 0,
             "VALOR_CONCILIADO_R$": r04["kpis"]["valor_em_risco"]},
            {"MÓDULO": "M05 – CB Debitados",
             "TOTAL": r05["kpis"]["total"],
             "CONCILIADOS": r05["kpis"]["conciliados"],
             "DIVERGENTES": r05["kpis"]["divergentes"],
             "SEM_PAR": r05["kpis"]["pendentes"],
             "TAXA_CONCILIAÇÃO_%": r05["kpis"]["taxa_conciliacao_pct"],
             "VALOR_CONCILIADO_R$": r05["kpis"]["valor_debitado"]},
            {"MÓDULO": "M06 – Taxas e Custos",
             "TOTAL": int(r06["kpis"]["tpv"]),
             "CONCILIADOS": int(r06["kpis"]["tpv"] - r06["kpis"]["diff_mdr"]),
             "DIVERGENTES": int(r06["kpis"]["n_divergencias_mdr"]),
             "SEM_PAR": 0,
             "TAXA_CONCILIAÇÃO_%": 100.0 - r06["kpis"]["custo_pct_tpv"],
             "VALOR_CONCILIADO_R$": r06["kpis"]["tpv"]},
            {"MÓDULO": "M07 – Fluxo de Caixa",
             "TOTAL": r07["kpis"]["dias_negativo"] + 90,
             "CONCILIADOS": 90 - r07["kpis"]["dias_negativo"],
             "DIVERGENTES": r07["kpis"]["dias_negativo"],
             "SEM_PAR": 0,
             "TAXA_CONCILIAÇÃO_%": round((90-r07["kpis"]["dias_negativo"])/90*100,2),
             "VALOR_CONCILIADO_R$": r07["kpis"]["saldo_final"]},
            {"MÓDULO": "M08 – Parcelamento",
             "TOTAL": r08["kpis"]["total"],
             "CONCILIADOS": r08["kpis"]["conciliados"],
             "DIVERGENTES": r08["kpis"]["divergentes"],
             "SEM_PAR": r08["kpis"].get("sem_par", 0),
             "TAXA_CONCILIAÇÃO_%": r08["kpis"]["taxa_conciliacao_pct"],
             "VALOR_CONCILIADO_R$": 0},
            {"MÓDULO": "M09 – Antecipações",
             "TOTAL": len(r09["resultado"]),
             "CONCILIADOS": len(r09["resultado"]),
             "DIVERGENTES": 0,
             "SEM_PAR": 0,
             "TAXA_CONCILIAÇÃO_%": 100.0,
             "VALOR_CONCILIADO_R$": r09["kpis"]["valor_liquido_antecipado"]},
            {"MÓDULO": "M10 – Recebíveis Líquidos",
             "TOTAL": len(r10["resultado"]),
             "CONCILIADOS": int((r10["resultado"]["STATUS_CONCILIACAO"]=="CONCILIADO").sum()),
             "DIVERGENTES": 0,
             "SEM_PAR": int((r10["resultado"]["STATUS_CONCILIACAO"]=="ANOMALIA").sum()),
             "TAXA_CONCILIAÇÃO_%": round(int((r10["resultado"]["STATUS_CONCILIACAO"]=="CONCILIADO").sum())/len(r10["resultado"])*100,2) if len(r10["resultado"]) else 0,
             "VALOR_CONCILIADO_R$": r10["kpis"]["liquido"]},
            {"MÓDULO": "M11 – Remessas para Comércios",
             "TOTAL": len(r11["resultado"]),
             "CONCILIADOS": len(r11["resultado"]),
             "DIVERGENTES": 0,
             "SEM_PAR": 0,
             "TAXA_CONCILIAÇÃO_%": 100.0,
             "VALOR_CONCILIADO_R$": r11["kpis"]["total_remessas"]},
        ])
        df_sumario.to_excel(writer, sheet_name="00_SUMÁRIO", index=False)

        # ── Módulos individuais ───────────────────────────────────────────────
        sheets = [
            ("01_PAGAMENTOS",    r01["resultado"]),
            ("02_REEMBOLSOS",    r02["resultado"]),
            ("03_CB_NOTIFICADOS",r03["resultado"]),
            ("04_DISPUTAS",      r04["resultado"]),
            ("05_CB_DEBITADOS",  r05["resultado"]),
            ("06_TAXAS_CUSTOS",  r06["resultado"]),
            ("07_FLUXO_CAIXA",   r07["resultado"]),
            ("08_PARCELAMENTO",  r08["resultado"]),
            ("09_ANTECIPACOES",  r09["resultado"]),
            ("10_RECEBIVEIS",    r10["resultado"]),
            ("11_REMESSAS",      r11["resultado"]),
        ]
        for sheet_name, df in sheets:
            df_out = df.copy()
            for col in df_out.select_dtypes(include=["datetimetz"]).columns:
                df_out[col] = df_out[col].dt.strftime("%Y-%m-%d %H:%M")
            df_out.head(30000).to_excel(writer, sheet_name=sheet_name, index=False)

        # Extras
        def strip_tz(df):
            df = df.copy()
            for col in df.select_dtypes(include=["datetimetz"]).columns:
                df[col] = df[col].dt.tz_localize(None)
            return df
        strip_tz(r06["resumo_bandeira"]).to_excel(writer, sheet_name="06_MDR_RESUMO",   index=False)
        strip_tz(r04["pipeline"]).to_excel(writer,        sheet_name="04_PIPELINE_CB",  index=False)
        strip_tz(r10["agenda"]).to_excel(writer,           sheet_name="10_AGENDA",       index=False)
        strip_tz(r11["top10_merchants"]).to_excel(writer,  sheet_name="11_TOP_MERCHANTS",index=False)

        # Dicionário de custos
        get_df_mdr().to_excel(writer,        sheet_name="MDR_DICIONARIO",    index=False)
        get_df_tarifas().to_excel(writer,    sheet_name="MDR_TARIFAS",       index=False)
        get_df_impostos().to_excel(writer,   sheet_name="MDR_IMPOSTOS",      index=False)
        get_df_cb_reasons().to_excel(writer, sheet_name="CB_REASON_CODES",   index=False)

    # ── Pós-formatação com openpyxl ───────────────────────────────────────────
    if OPENPYXL_OK:
        wb = load_workbook(OUTPUT_PATH)

        for ws_name in wb.sheetnames:
            ws = wb[ws_name]
            # Freeze top row
            ws.freeze_panes = "A2"
            # Estilizar header row 1
            if ws.max_row >= 1:
                for cell in ws[1]:
                    cell.fill      = fill(C["dark"])
                    cell.font      = font(bold=True, color=C["white"], size=10)
                    cell.alignment = align("center")
                    cell.border    = thin_border()
            # Colorir coluna STATUS_CONCILIACAO
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    if cell.column_letter and ws.cell(1, cell.column).value == "STATUS_CONCILIACAO":
                        val = str(cell.value) if cell.value else ""
                        style_status_cell(cell, val)
                    if cell.row % 2 == 0 and cell.fill.fgColor.rgb in ["00000000","FFFFFFFF"]:
                        cell.fill = fill(C["gray"])

        # Aba sumário: destaque especial
        ws_sum = wb["00_SUMÁRIO"]
        ws_sum.freeze_panes = "A2"
        for row in ws_sum.iter_rows(min_row=2):
            taxa_cell = None
            for cell in row:
                hdr = ws_sum.cell(1, cell.column).value
                if hdr == "TAXA_CONCILIAÇÃO_%":
                    taxa_cell = cell
                    try:
                        val = float(cell.value or 0)
                        if val >= 95: cell.fill = fill(C["green"]); cell.font = font(bold=True, color=C["acc"])
                        elif val >= 85: cell.fill = fill(C["yel_bg"]); cell.font = font(bold=True, color="7B4F00")
                        else: cell.fill = fill(C["red_bg"]); cell.font = font(bold=True, color=C["red"])
                    except: pass
                if hdr == "VALOR_CONCILIADO_R$":
                    try:
                        cell.number_format = 'R$ #,##0.00'
                    except: pass

        wb.save(OUTPUT_PATH)
        log.info(f"Formatação aplicada com sucesso.")

    log.info("=" * 70)
    log.info(f"PIPELINE CONCLUÍDO: {OUTPUT_PATH}")
    log.info("=" * 70)
    return all_kpis


if __name__ == "__main__":
    kpis = run_pipeline()
    print("\n=== RESUMO FINAL DE KPIs ===")
    for mod, k in kpis.items():
        print(f"{mod}: {k}")
